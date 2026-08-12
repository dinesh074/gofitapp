"""
One-off script: extracts genuinely Indian-relevant dishes from the UK's
official COFID 2021 dataset (McCance & Widdowson's Composition of Foods
Integrated Dataset, Public Health England) and inserts the ones NOT already
in our `foods` table, using the same health_score/benefits_watchouts formula
as build_db_v2.py for consistency, and main.py's classify_diet_tags for
jain/sattvic status.

Source file: cofid_temp.xlsx (downloaded from
https://www.gov.uk/government/publications/composition-of-foods-integrated-dataset-cofid)

Real government analytical/recipe-survey data -- many entries are explicitly
labelled by community (Punjabi/Gujarati/Bangladeshi dish, dietary survey
records), not estimates. Only inserts NEW keys; never touches existing rows.
"""
import os
import re
import json
import sys

import openpyxl
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
import psycopg  # noqa: E402

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from main import classify_diet_tags  # noqa: E402

# Copied from build_db_v2.py rather than imported: that script has no
# __main__ guard and regenerates indian_food_db.json as a module-level side
# effect just from being imported (confirmed harmless/idempotent here, but
# not something to rely on twice).


def health_score(kcal100, protein100, fibre100, sfa_mg100, sugar100, sodium_mg100):
    if not kcal100 or kcal100 <= 0:
        return None
    score = 50.0
    score += min(20.0, (protein100 / kcal100 * 100.0) * 1.0)
    if fibre100 is not None:
        score += min(15.0, fibre100 * 2.0)
    if sfa_mg100 is not None:
        score -= min(20.0, (sfa_mg100 / 1000.0) * 1.5)
    if sugar100 is not None:
        score -= min(20.0, sugar100 * 1.0)
    if sodium_mg100 is not None:
        score -= min(15.0, max(0.0, (sodium_mg100 - 400.0) / 100.0))
    return round(max(0.0, min(100.0, score)), 1)


def benefits_watchouts(protein, fibre, iron, calcium, fat, sfa_mg, sugar, sodium_mg, kcal):
    b, w = [], []
    if protein is not None:
        if protein >= 15:
            b.append("High protein")
        elif protein >= 8:
            b.append("Good source of protein")
    if fibre is not None:
        if fibre >= 5:
            b.append("High in fiber")
        elif fibre >= 3:
            b.append("Good source of fiber")
    if iron is not None and iron >= 3:
        b.append("Good source of iron")
    if calcium is not None and calcium >= 200:
        b.append("Good source of calcium")
    if fat is not None and fat <= 3 and (kcal or 0) < 300:
        b.append("Low fat")
    if sfa_mg is not None and sfa_mg >= 5000:
        w.append("High in saturated fat")
    if sugar is not None and sugar >= 15:
        w.append("High in added sugar")
    if sodium_mg is not None and sodium_mg >= 600:
        w.append("High in sodium")
    if fat is not None and fat >= 20:
        w.append("High in fat")
    if kcal is not None and kcal >= 500:
        w.append("Calorie-dense - mind your portion")
    return b, w

WB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cofid_temp.xlsx")

KEYWORDS = [
    "curry", "chapati", "roti", "paratha", "dal ", "dal,", "dhal", "paneer",
    "samosa", "biryani", "biriyani", "tikka", "naan", "poppadom", "papad",
    "tandoori", "korma", "masala", "pakora", "bhajia", "bhaji", "dosa",
    "idli", "raita", "vindaloo", "dopiaza", "jalfrezi", "balti", "madras",
    "dhansak", "rogan josh", "pilau", "pulao", "chick pea", "chickpea",
    "channa", "chana", "gulab jamun", "barfi", "halwa", "kheer", "lassi",
    "chutney, mango", "bombay",
]


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s).lower()).strip("_")


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "N", "Tr", "tr", "-"):
        return 0.0 if s in ("Tr", "tr") else None
    try:
        return float(s)
    except ValueError:
        return None


def main():
    wb = openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
    prox = wb["1.3 Proximates"]
    inorg = wb["1.4 Inorganics"]

    # index inorganics by food code
    inorg_by_code = {}
    for row in inorg.iter_rows(min_row=4, values_only=True):
        code = row[0]
        if code:
            inorg_by_code[code] = row

    existing_keys = set(json.load(open(os.path.join(os.path.dirname(__file__), "existing_keys.json"))))

    candidates = []
    # Word-boundary matching -- a naive substring check on KEYWORDS matched
    # "masala" inside "tara-MASALA-ta" (a Greek dish), which is not Indian
    # food at all. \b keeps this from firing on prefix/suffix overlaps like
    # that instead of a real standalone word.
    # Leading \b only (not trailing) -- keeps plurals like "samosas",
    # "pakoras", "curries" matching (a trailing \b would reject those, since
    # there's a word character, not a boundary, right after the stem) while
    # still rejecting "masala" appearing mid-word as in "tara-MASALA-ta"
    # (Greek, not Indian) -- there's genuinely no word boundary immediately
    # before the "m" there, so \b correctly excludes it.
    kw_pattern = re.compile(
        r"\b(" + "|".join(re.escape(k.strip(" ,")) for k in KEYWORDS) + r")"
    )
    for row in prox.iter_rows(min_row=4, values_only=True):
        code, name, desc = row[0], row[1], row[2]
        if not name:
            continue
        hay = f"{name} {desc or ''}".lower()
        if not kw_pattern.search(hay):
            continue

        kcal = to_float(row[12])
        protein = to_float(row[9])
        carbs = to_float(row[11])
        fat = to_float(row[10])
        sugar = to_float(row[16])
        fiber = to_float(row[25])  # AOAC fibre
        if kcal is None or protein is None or carbs is None or fat is None:
            continue  # can't trust an entry missing core macros

        inorg_row = inorg_by_code.get(code)
        sodium = to_float(inorg_row[7]) if inorg_row else None
        potassium = to_float(inorg_row[8]) if inorg_row else None
        calcium = to_float(inorg_row[9]) if inorg_row else None
        iron = to_float(inorg_row[12]) if inorg_row else None

        key = norm(name)
        if not key or key in existing_keys:
            continue

        candidates.append(
            {
                "key": key,
                "name": name,
                "desc": desc,
                "kcal": kcal, "protein": protein, "carbs": carbs, "fat": fat,
                "fiber": fiber, "sugar": sugar,
                "sodium": sodium, "potassium": potassium, "calcium": calcium, "iron": iron,
            }
        )
        existing_keys.add(key)  # de-dup within this batch too

    print(f"Found {len(candidates)} new Indian-relevant dishes not already in our DB.")

    if "--dry-run" in sys.argv:
        for c in candidates:
            print(f"  {c['key']:40s} kcal={c['kcal']:.0f} P={c['protein']:.1f} C={c['carbs']:.1f} "
                  f"F={c['fat']:.1f} fib={c['fiber']} sug={c['sugar']} Na={c['sodium']} | {c['name']}")
        return

    url = os.environ["DATABASE_URL"]
    inserted = 0
    with psycopg.connect(url) as conn:
        with conn.cursor() as cur:
            cur.execute("set search_path to gofit")
            for c in candidates:
                score = health_score(c["kcal"], c["protein"], c["fiber"], None, c["sugar"], c["sodium"])
                benefits, watchouts = benefits_watchouts(
                    c["protein"], c["fiber"], c["iron"], c["calcium"], c["fat"],
                    None, c["sugar"], c["sodium"], c["kcal"],
                )
                jain_status, sattvic_status = classify_diet_tags(c["name"], [c["name"]])
                micros = {}
                if c["potassium"] is not None:
                    micros["potassium_mg"] = c["potassium"]
                cur.execute(
                    """
                    INSERT INTO foods
                        (key, unit, kcal_per_unit, protein_g, carbs_g, fat_g, fiber_g, sugar_g,
                         sodium_mg, potassium_mg, calcium_mg, iron_mg, health_score,
                         benefits_json, watch_outs_json, micros_json, aliases_json,
                         source_name, source, jain_status, sattvic_status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (key) DO NOTHING
                    """,
                    (
                        c["key"], "100g", c["kcal"], c["protein"], c["carbs"], c["fat"],
                        c["fiber"], c["sugar"], c["sodium"], c["potassium"], c["calcium"], c["iron"],
                        score, json.dumps(benefits), json.dumps(watchouts), json.dumps(micros),
                        json.dumps([c["name"]]),
                        "COFID 2021 (UK Public Health England / McCance & Widdowson's)",
                        "government_uk", jain_status, sattvic_status,
                    ),
                )
                inserted += 1
        conn.commit()
    print(f"Inserted {inserted} new rows into foods.")


if __name__ == "__main__":
    main()
